import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox
import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
import threading
from moviepy.editor import VideoFileClip
import humanize
import subprocess
import json

class FileInfoExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("File Info Extractor")
        self.root.geometry("432x150")

        self.folder_path = tk.StringVar()
        self.file_info_list = []

        # ====== 資料夾選擇區塊 ======
        label = tk.Label(root, text="Select Root Folder:", anchor='center')
        label.pack(fill='x', pady=4)
        frame = tk.Frame(root)
        frame.pack(pady=2)
        tk.Entry(frame, textvariable=self.folder_path, width=45).pack(side=tk.LEFT, padx=2)
        tk.Button(frame, text="Browse", command=self.browse_folder, width=8).pack(side=tk.LEFT)

        # ====== 進度條區塊 ======
        self.progress = ttk.Progressbar(root, orient='horizontal', length=350, mode='determinate')
        self.progress.pack(pady=2)
        self.progress_label = tk.Label(root, text="", anchor='center')
        self.progress_label.pack(fill='x')

        # ====== 操作按鈕區塊 ======
        btn_frame = tk.Frame(root)
        btn_frame.pack(pady=8)
        btn_width = 16
        btn_height = 1
        tk.Button(btn_frame, text="Scan", width=btn_width, height=btn_height, command=self.scan_files_thread).pack(side=tk.LEFT, padx=12)
        tk.Button(btn_frame, text="Export to Excel", width=btn_width, height=btn_height, command=self.export_excel).pack(side=tk.LEFT, padx=12)

    def browse_folder(self):
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            self.folder_path.set(folder_selected)

    def format_duration(self, seconds):
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        seconds = int(seconds % 60)
        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"

    def format_size(self, size_bytes):
        if size_bytes >= 1024**3:
            return f"{size_bytes/1024**3:.2f} GB"
        else:
            return f"{size_bytes/1024**2:.2f} MB"

    # === 新增功能：使用 ffprobe 獲取影片編碼 (更可靠的方法) ===
    def get_video_codec(self, file_path):
        """使用 ffprobe 直接獲取影片編碼，結果更準確"""
        command = [
            'ffprobe',
            '-v', 'quiet',
            '-print_format', 'json',
            '-select_streams', 'v:0',
            '-show_entries', 'stream=codec_name',
            file_path
        ]
        try:
            # 執行命令，並設定 timeout 以免卡在損毀的檔案
            result = subprocess.run(command, capture_output=True, text=True, check=True, timeout=30)
            data = json.loads(result.stdout)
            if data and 'streams' in data and data['streams']:
                codec = data['streams'][0].get('codec_name', 'N/A').upper()
                if codec == 'HEVC':
                    return 'H265'
                return codec
            return 'N/A'
        except FileNotFoundError:
            # 如果系統找不到 ffprobe 命令，則返回特定錯誤訊息
            messagebox.showerror("錯誤", "找不到 ffprobe 命令。\n請確認您已安裝 FFmpeg 並將其路徑加入系統環境變數中。")
            self.root.quit() # 終止程式
            return 'FFPROBE_NOT_FOUND'
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired, json.JSONDecodeError, IndexError):
            # 處理各種可能的錯誤：ffprobe執行失敗、超時、JSON解析失敗、沒有影片串流等
            return 'N/A'

    def scan_files_thread(self):
        t = threading.Thread(target=self.scan_files)
        t.daemon = True
        t.start()

    def scan_files(self):
        root_folder = self.folder_path.get()
        if not root_folder:
            messagebox.showwarning("Warning", "Please select a folder first!")
            return

        self.file_info_list = []
        video_extensions = ('.mp4', '.ts', '.mkv', '.avi', '.mov')

        total_files = 0
        for root, dirs, files in os.walk(root_folder):
            for file in files:
                if file.lower().endswith(video_extensions):
                    total_files += 1
        if total_files == 0:
            self.progress['value'] = 0
            self.progress_label.config(text="無影片檔案")
            return
        self.progress['maximum'] = total_files
        self.progress['value'] = 0
        self.progress_label.config(text="0%")
        self.root.update_idletasks()

        processed = 0
        try:
            for root, dirs, files in os.walk(root_folder):
                for file in files:
                    if file.lower().endswith(video_extensions):
                        file_path = os.path.join(root, file)
                        
                        # === 修改處：呼叫新的 get_video_codec 函式 ===
                        codec = self.get_video_codec(file_path)
                        if codec == 'FFPROBE_NOT_FOUND': # 如果找不到 ffprobe，則停止掃描
                            return

                        try:
                            file_size_bytes = os.path.getsize(file_path)
                            with VideoFileClip(file_path) as clip:
                                duration_seconds = clip.duration
                                resolution_str = f"{clip.size[0]}x{clip.size[1]}"
                            
                            if duration_seconds >= 3600:
                                avg_size_per_hour = file_size_bytes / (duration_seconds / 3600)
                                avg_size_str = f"{avg_size_per_hour/1024**3:.2f} GB/hr"
                            else:
                                avg_size_str = '無法計算'
                            
                            file_info = {
                                'File Name': file,
                                'Codec': codec,
                                'Resolution': resolution_str,
                                'Video Duration': self.format_duration(duration_seconds),
                                'File Size': self.format_size(file_size_bytes),
                                'File Size in Bytes': file_size_bytes,
                                'Average File Size Per Hour': avg_size_str
                            }
                            self.file_info_list.append(file_info)
                        except Exception as e:
                            print(f"Error processing file with moviepy {file_path}: {str(e)}")
                            continue
                        
                        processed += 1
                        self.progress['value'] = processed
                        percent = int(processed / total_files * 100)
                        self.progress_label.config(text=f"{percent}%")
                        self.root.update_idletasks()
            self.progress_label.config(text="掃描完成！")
            messagebox.showinfo("Info", f"Scan complete! {len(self.file_info_list)} files found.")
        except Exception as e:
            self.progress_label.config(text="掃描失敗")
            messagebox.showerror("Error", f"An error occurred during scanning: {str(e)}")

    def export_excel(self):
        if not self.file_info_list:
            messagebox.showwarning("Warning", "No data to export. Please scan first.")
            return
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Save as Excel file"
        )
        if not save_path:
            return
        
        df = pd.DataFrame(self.file_info_list)
        
        column_order = [
            'File Name', 'Codec', 'Resolution', 'Video Duration', 
            'File Size', 'File Size in Bytes', 'Average File Size Per Hour'
        ]
        df = df[column_order]
        
        df.to_excel(save_path, index=False)

        wb = load_workbook(save_path)
        ws = wb.active
        a1_length = len(str(ws['A1'].value))
        ws.column_dimensions['A'].width = a1_length + 2
        for col in ws.iter_cols(min_col=2, max_col=ws.max_column):
            col_letter = col[0].column_letter
            max_length = len(str(col[0].value))
            ws.column_dimensions[col_letter].width = max_length + 2
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='left', vertical='center')
        wb.save(save_path)
        messagebox.showinfo("Success", f"Exported to:\n{save_path}")

if __name__ == "__main__":
    root = tk.Tk()
    app = FileInfoExtractorApp(root)
    root.mainloop()
